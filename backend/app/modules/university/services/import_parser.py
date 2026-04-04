from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import load_workbook

# Ключи — заголовок после normalize_header (нижний регистр, пробелы схлопнуты, _ → пробел).
_HEADER_ALIASES: dict[str, str] = {
  "фио": "full_name",
  "fio": "full_name",
  "full name": "full_name",
  "ф.и.о.": "full_name",
  "год": "year",
  "year": "year",
  "год выпуска": "year",
  "специальность": "specialty",
  "specialty": "specialty",
  "спец": "specialty",
  "номер диплома": "diploma_number",
  "diploma number": "diploma_number",
  "номер": "diploma_number",
  "diploma": "diploma_number",
  "№ диплома": "diploma_number",
}


def normalize_header(value: str | None) -> str:
  if value is None:
    return ""
  text = str(value).strip().lower().replace("_", " ")
  text = re.sub(r"\s+", " ", text)
  return text


def map_header(raw: str) -> str | None:
  key = normalize_header(raw)
  return _HEADER_ALIASES.get(key)


def parse_csv_bytes(data: bytes) -> list[dict[str, str]]:
  text = data.decode("utf-8-sig")
  reader = csv.DictReader(io.StringIO(text))
  if not reader.fieldnames:
    return []
  header_map = {name: map_header(name) for name in reader.fieldnames if name}
  rows: list[dict[str, str]] = []
  for row in reader:
    normalized: dict[str, str] = {}
    for raw_key, cell in row.items():
      logical = header_map.get(raw_key or "")
      if logical is None:
        continue
      if cell is None:
        normalized[logical] = ""
      else:
        normalized[logical] = str(cell).strip()
    rows.append(normalized)
  return rows


def parse_xlsx_bytes(data: bytes) -> list[dict[str, str]]:
  workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
  try:
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if header_row is None:
      return []
    headers: list[str | None] = [str(c).strip() if c is not None else None for c in header_row]
    logical_headers: list[str | None] = [map_header(h) for h in headers]
    rows: list[dict[str, str]] = []
    for data_row in row_iter:
      if data_row is None:
        continue
      if all(v is None or str(v).strip() == "" for v in data_row):
        continue
      normalized: dict[str, str] = {}
      for col_idx, logical in enumerate(logical_headers):
        if logical is None:
          continue
        raw_cell: Any = data_row[col_idx] if col_idx < len(data_row) else None
        if raw_cell is None:
          normalized[logical] = ""
        else:
          normalized[logical] = str(raw_cell).strip()
      rows.append(normalized)
    return rows
  finally:
    workbook.close()
